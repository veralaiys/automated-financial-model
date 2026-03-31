from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, Series
from model import load_data, calc_income_metrics, calc_balance_metrics
from model import calc_cashflow_metrics, calc_dcf


# Styles
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SUBHEAD_FILL = PatternFill("solid", fgColor="2E75B6")
ALT_ROW_FILL = PatternFill("solid", fgColor="EBF3FB")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEAD_FONT = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(size=10)
CENTER = Alignment(horizontal='center')
RIGHT = Alignment(horizontal="right")


def format_sheet(ws, df, currency_cols=None, pct_cols=None):
    ws.column_dimensions['A'].width = 22
    for col in ws.iter_cols(min_col=2, max_col=ws.max_column):
        ws.column_dimensions[col[0].column_letter].width = 18
    
    for i, row in enumerate(ws.iter_rows()):
        for j, cell in enumerate(row):
            if i == 0:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = CENTER
            elif j == 0:
                cell.fill = SUBHEAD_FILL
                cell.font = SUBHEAD_FONT
                cell.alignment = CENTER
            else:
                cell.font = BODY_FONT
                cell.alignment = RIGHT
                if i % 2 == 0:
                    cell.fill = ALT_ROW_FILL
    
    for col_idx, col_name in enumerate(df.columns, start=2):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        for row_idx in range(2, ws.max_row + 1):
            cell = ws[f"{col_letter}{row_idx}"]
            if currency_cols and col_name in currency_cols:
                cell.number_format = '$#,##0.0,, "B"'
            elif pct_cols and col_name in pct_cols:
                cell.number_format = '0.00"%"'



def build_workbook(income_metrics, balance_metrics, cf_metrics, dcf_table, intrinsic_value):
    wb = Workbook()
    wb.remove(wb.active)

    # Summary sheet
    ws_sum = wb.create_sheet("Summary", 0)
    ws_sum['B2'] = 'Apple Inc. — Financial Model Summary'
    ws_sum['B2'].font = Font(bold=True, size=14, color="1F4E79")

    ws_sum['B4'] = "Metric"
    ws_sum['C4'] = "FY2025"
    ws_sum['D4'] = "FY2024"
    for cell in [ws_sum['B4'], ws_sum['C4'], ws_sum['D4']]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    summary_rows = {
        ("Revenue", income_metrics.loc['2025-09-30', 'Revenue'], income_metrics.loc['2024-09-30', 'Revenue']),
        ("Gross Margin %", income_metrics.loc['2025-09-30', 'Gross Margin %'], income_metrics.loc['2024-09-30', 'Gross Margin %']),
        ("EBIT Margin %", income_metrics.loc['2025-09-30', 'EBIT Margin %'], income_metrics.loc['2024-09-30', 'EBIT Margin %']),
        ("Net Income", income_metrics.loc['2025-09-30', 'Net Income'], income_metrics.loc['2024-09-30', 'Net Income']),
        ("Free Cash Flow", cf_metrics.loc['2025-09-30', 'Free Cash Flow'], cf_metrics.loc['2024-09-30', 'Free Cash Flow']),
        ("Total Debt", balance_metrics.loc['2025-09-30', 'Total Debt'], balance_metrics.loc['2024-09-30', 'Total Debt']),
        ("Net Debt", balance_metrics.loc['2025-09-30', 'Net Debt'], balance_metrics.loc['2024-09-30', 'Net Debt']),
        ("DCF Intrinsic Value ($B)", intrinsic_value, None)
    }

    pct_rows =  {"Gross Margin %", "EBIT Margin %"}

    for i, (label, val1, val2) in enumerate(summary_rows, start=5):
        ws_sum[f'B{i}'] = label
        ws_sum[f'B{i}'].font = Font(bold=True, size=10)
        ws_sum[f'C{i}'] = val1
        ws_sum[f'D{i}'] = val2 if val2 is not None else "—"

        if label == "DCF Intrinsic Value ($B)":
            fmt = '"$"#,##0.00"B"'
        elif label in pct_rows:
            fmt = '0.00"%"'
        else:
            fmt = '$#,##0.0,,"B"'

        ws_sum[f'C{i}'].number_format = fmt
        if val2:
            ws_sum[f'D{i}'].number_format = fmt
        if i % 2 == 0:
            for col in ['B', 'C', 'D']:
                ws_sum[f'{col}{i}'].fill = ALT_ROW_FILL

    ws_sum.column_dimensions['B'].width = 28
    ws_sum.column_dimensions['C'].width = 26
    ws_sum.column_dimensions['D'].width = 16

    # Income statement, Balance, Cash flow, and DCF valuation sheets
    sheets = [
        ("Income Statement", income_metrics,
         ['Revenue', 'Gross Profit', 'EBIT', 'EBITDA', 'Net Income'],
         ['Gross Margin %', 'EBIT Margin %', 'Net Margin %']),
        ("Balance Sheet", balance_metrics,
         ['Total Assets', 'Total Liabilities', 'Total Equity', 'Total Debt', 'Cash', 'Net Debt'],
         None),
        ("Cash Flow", cf_metrics,
         ['Operating CF', 'Capex', 'Free Cash Flow', 'D&A'],
         None),
        ("DCF Valuation", dcf_table, None, None)
    ]

    for sheet_name, df, curr_cols, pct_cols in sheets:
        ws = wb.create_sheet(sheet_name)
        for row in dataframe_to_rows(df, index=True, header=True):
            ws.append(row)
        ws.delete_rows(2)
        format_sheet(ws, df, currency_cols=curr_cols, pct_cols=pct_cols)

        # DCF values are already in Billions
        if sheet_name == "DCF Valuation":
            for row in ws.iter_rows(min_row=2, min_col=2, max_col=ws.max_column):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '"$"#,##0.00"B"'

    # Chart
    ws_chart = wb.create_sheet("Chart")
    years = list(income_metrics.index)
    revenues = list(income_metrics['Revenue'])
    fcfs = list (cf_metrics['Free Cash Flow'])

    ws_chart['A1'] = "Year"
    ws_chart['B1'] = "Revenue ($B)"
    ws_chart['C1'] = "Free Cash Flow ($B)"

    for i, (yr, rev, fcf) in enumerate(zip(years, revenues, fcfs), start=2):
        ws_chart[f'A{i}'] = str(yr[:4])
        ws_chart[f'B{i}'] = round(float(rev) / 1e9, 1)
        ws_chart[f'C{i}'] = round(float(fcf) / 1e9, 1)
        
    chart = BarChart()
    chart.type = "col"
    chart.title = "Apple - Revenue vs Free Cash Flow"
    chart.y_axis.title = "USD Billions"
    chart.x_axis.title = "Fiscal Year"
    chart.style = 10
    chart.width = 20
    chart.height = 12
    chart.grouping = 'clustered'

    data = Reference(ws_chart, min_col=2, max_col=3, min_row=1, max_row=len(years)+1)
    cats = Reference(ws_chart, min_col=1, min_row=2, max_row=len(years)+1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws_chart.add_chart(chart, "E2")

    wb.save("financial_model.xlsx")
    print("financial_model.xlsx saved!")


if __name__ == "__main__":
    income_stmt, balance_sheet, cash_flow = load_data()
    income_metrics = calc_income_metrics(income_stmt)
    balance_metrics = calc_balance_metrics(balance_sheet)
    cf_metrics = calc_cashflow_metrics(cash_flow)
    dcf_table, intrinsic_value = calc_dcf(cf_metrics)
    build_workbook(income_metrics, balance_metrics, cf_metrics,dcf_table, intrinsic_value)